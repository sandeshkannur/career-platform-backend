import io
from typing import Dict, List, Tuple, Set
from sqlalchemy.orm import Session
from sqlalchemy import tuple_
from openpyxl import load_workbook

from app.models import Skill, KeySkill, SkillKeySkillMap


REQUIRED_SHEET_NORM = "key skills students skills"
COL_KEYSKILL = "Key Skill"
COL_STUDENT_SKILLS = "Student Skill(s) Mapped"


def _norm(val: str) -> str:
    """
    Conservative normalization for matching:
    - lower + trim
    - collapse whitespace
    """
    if val is None:
        return ""
    s = str(val).strip().lower()
    s = " ".join(s.split())
    return s


def _header_map(header_row: List[str]) -> Dict[str, int]:
    """
    Map column name -> index from the first row.
    """
    hm: Dict[str, int] = {}
    for idx, v in enumerate(header_row):
        if v is None:
            continue
        hm[str(v).strip()] = idx
    return hm

def _norm_sheet_name(name: str) -> str:
    # Lowercase and keep only alphanumerics/spaces, collapse whitespace
    s = "".join(ch.lower() if ch.isalnum() else " " for ch in name)
    return " ".join(s.split())

def ingest_skill_keyskill_map(db: Session, file_bytes: bytes, dry_run: bool = True) -> Dict:
    """
    Dry-run ingestion for PR46:
    - Parses Excel sheet: 'Key skills- Students skills'
    - Reads columns: 'Key Skill' and 'Student Skill(s) Mapped'
    - Resolves StudentSkill -> skills.id (canonical only)
    - Resolves KeySkill -> keyskills.id by normalized name
      - If multiple ids match the same normalized name => ambiguous orphan, skip
    - Inserts into DB only when dry_run=False
    - Returns the required report payload
    """

    # Load workbook from bytes
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    # Find sheet by normalized name (tolerant to hyphens/case/spaces)
    target_sheet = None
    for sn in wb.sheetnames:
        if _norm_sheet_name(sn) == REQUIRED_SHEET_NORM:
            target_sheet = sn
            break

    if not target_sheet:
        raise ValueError(
            f"Missing required sheet (expected something like 'Key skills- Students skills'). "
            f"Found sheets: {wb.sheetnames}"
        )

    ws = wb[target_sheet]

    # Build lookup: canonical skills (48)
    skills = db.query(Skill.id, Skill.name).all()
    skill_by_norm: Dict[str, int] = {_norm(name): sid for sid, name in skills}

    # Build lookup: keyskills (note: you have 1077)
    keyskills = db.query(KeySkill.id, KeySkill.name).all()
    keyskill_ids_by_norm: Dict[str, List[int]] = {}
    for kid, name in keyskills:
        nk = _norm(name)
        keyskill_ids_by_norm.setdefault(nk, []).append(kid)

    # Read header row (row 1)
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hm = _header_map(header)

    if COL_KEYSKILL not in hm or COL_STUDENT_SKILLS not in hm:
        raise ValueError(
            f"Missing required columns in sheet '{target_sheet}'. "
            f"Need '{COL_KEYSKILL}' and '{COL_STUDENT_SKILLS}'. Found: {list(hm.keys())}"
        )

    rows_received = 0
    rows_valid = 0

    inserted = 0

    # Orphan buckets (as required)
    student_skill_not_found: Set[str] = set()
    keyskill_not_found: Set[str] = set()
    ambiguous_keyskill: Set[str] = set()

    # Collect resolved (skill_id, keyskill_id) pairs
    resolved_pairs: Set[Tuple[int, int]] = set()

    # Iterate rows starting row 2
    for row in ws.iter_rows(min_row=2):
        keyskill_val = row[hm[COL_KEYSKILL]].value
        student_skills_val = row[hm[COL_STUDENT_SKILLS]].value

        # Skip fully blank rows
        if keyskill_val is None and student_skills_val is None:
            continue

        rows_received += 1

        # Resolve keyskill
        keyskill_norm = _norm(keyskill_val)
        if not keyskill_norm:
            keyskill_not_found.add(str(keyskill_val).strip())
            continue

        ks_ids = keyskill_ids_by_norm.get(keyskill_norm, [])
        if len(ks_ids) == 0:
            keyskill_not_found.add(str(keyskill_val).strip())
            continue
        if len(ks_ids) > 1:
            # IMPORTANT: do not guess
            ambiguous_keyskill.add(keyskill_norm)
            continue

        keyskill_id = ks_ids[0]

        # Resolve student skills (may be comma-separated / newline separated)
        raw = "" if student_skills_val is None else str(student_skills_val)
        parts = [p.strip() for p in raw.replace("\n", ",").split(",") if p.strip()]
        if not parts:
            student_skill_not_found.add(str(student_skills_val).strip())
            continue

        any_resolved = False
        for p in parts:
            sid = skill_by_norm.get(_norm(p))
            if not sid:
                # no alias guessing in PR46
                student_skill_not_found.add(p)
                continue

            any_resolved = True
            resolved_pairs.add((sid, keyskill_id))

        if any_resolved:
            rows_valid += 1

    # Count existing pairs (so response can report skipped_existing even in dry-run)
    # Count existing pairs (for reporting + idempotency)
    skipped_existing = 0
    existing = []
    if resolved_pairs:
        existing = (
            db.query(SkillKeySkillMap.skill_id, SkillKeySkillMap.keyskill_id)
            .filter(tuple_(SkillKeySkillMap.skill_id, SkillKeySkillMap.keyskill_id).in_(list(resolved_pairs)))
            .all()
        )
        skipped_existing = len(existing)

    # Commit mode: insert clean resolved pairs only (idempotent)
    if dry_run is False and resolved_pairs:
        existing_set = set(existing)
        # Insert only missing pairs; UNIQUE(skill_id, keyskill_id) prevents duplicates
        for (skill_id, keyskill_id) in sorted(resolved_pairs):
            if (skill_id, keyskill_id) in existing_set:
                continue

            db.add(
                SkillKeySkillMap(
                    skill_id=skill_id,
                    keyskill_id=keyskill_id,
                    weight=1.0,  # semantic presence ONLY (no scoring meaning)
                )
            )
            inserted += 1

        db.commit()
        # Insert only missing pairs; UNIQUE(skill_id, keyskill_id) prevents duplicates
        for (skill_id, keyskill_id) in sorted(resolved_pairs):
            exists = (
                db.query(SkillKeySkillMap.id)
                .filter(
                    SkillKeySkillMap.skill_id == skill_id,
                    SkillKeySkillMap.keyskill_id == keyskill_id,
                )
                .first()
            )
            if exists:
                continue

            db.add(
                SkillKeySkillMap(
                    skill_id=skill_id,
                    keyskill_id=keyskill_id,
                    weight=1.0,  # semantic presence ONLY (no scoring meaning)
                )
            )
            inserted += 1

        db.commit()

    return {
        "rows_received": rows_received,
        "rows_valid": rows_valid,
        "inserted": inserted,
        "skipped_existing": skipped_existing,
        "orphans": {
            "student_skill_not_found": sorted(student_skill_not_found),
            "keyskill_not_found": sorted(keyskill_not_found),
            "ambiguous_keyskill": sorted(ambiguous_keyskill),
        },
        "dry_run": dry_run
    }
